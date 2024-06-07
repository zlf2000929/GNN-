import torch
from torch_geometric.nn import GCNConv, MessagePassing
from torch_geometric.data import DataLoader
from torch_geometric.utils import add_self_loops, degree
import torch.nn.functional as F
import torch.nn.init as init
import math
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
class CustomGCNConv(GCNConv):
    def __init__(self, in_channels, out_channels, **kwargs):
        super().__init__(in_channels, out_channels, **kwargs)
        self.initialized = False
    def forward(self, x, edge_index, edge_weight=None):
        if not self.initialized:
            data_mean = x.mean()
            data_std = x.std()
            xavier_var = 2.0 / (self.in_channels + self.out_channels)
            custom_mean = torch.sqrt(data_mean * (xavier_var ** 2 / data_std))
            torch.nn.init.normal_(self.lin.weight, mean=custom_mean.item(), std=math.sqrt(xavier_var))
            if self.lin.bias is not None:
                torch.nn.init.constant_(self.lin.bias, 0)
            self.initialized = True
        return super().forward(x, edge_index, edge_weight)
class GCN(torch.nn.Module):
    def __init__(self, input_dim, dropout_rate1, dropout_rate2, dropout_rate3, dropout_rate4, dropout_rate5, dropout_rate6, dropout_rate7, dropout_rate8, dropout_rate9):
        super(GCN, self).__init__()
        self.input_dim = input_dim
        self.dropout_rate1 = dropout_rate1
        self.dropout_rate2 = dropout_rate2
        self.dropout_rate3 = dropout_rate3
        self.dropout_rate4 = dropout_rate4
        self.dropout_rate5 = dropout_rate5
        self.dropout_rate6 = dropout_rate6
        self.dropout_rate7 = dropout_rate7
        self.dropout_rate8 = dropout_rate8
        self.dropout_rate9 = dropout_rate9
        self.hidden1 = self.input_dim - self.input_dim % 16
        #首先定义三个通道
        self.GCN_L_1 = CustomGCNConv(self.input_dim, self.hidden1)
        self.GCN_L_1_bn = BatchNorm1d(self.hidden1)
        self.GAT_L_1 = GATConv(self.hidden1, self.hidden1)
        self.GAT_L_1_bn = BatchNorm1d(self.hidden1)
        self.GCN_M_1 = CustomGCNConv(self.input_dim, self.hidden1)
        self.GCN_M_1_bn = BatchNorm1d(self.hidden1)
        self.GCN_M_2 = CustomGCNConv(self.hidden1, self.hidden1 // 2)
        self.GCN_M_2_bn = BatchNorm1d(self.hidden1 // 2)
        self.GAT_M_2 = GATConv(self.hidden1 // 2, self.hidden1 // 2)
        self.GAT_M_2_bn = BatchNorm1d(self.hidden1 // 2)
        self.GCN_R_1 = CustomGCNConv(self.input_dim, self.hidden1)
        self.GCN_R_1_bn = BatchNorm1d(self.hidden1)
        self.GCN_R_2 = CustomGCNConv(self.hidden1, self.hidden1 // 2)
        self.GCN_R_2_bn = BatchNorm1d(self.hidden1 // 2)
        self.GCN_R_3 = CustomGCNConv(self.hidden1 // 2, self.hidden1 // 4)
        self.GCN_R_3_bn = BatchNorm1d(self.hidden1 // 4)
        self.GAT_R_3 = GATConv(self.hidden1 // 4, self.hidden1 // 4)
        self.GAT_R_3_bn = BatchNorm1d(self.hidden1 // 4)
        #卷积之后再拼接的部分
        self.GCN_GAT = GATConv(self.hidden1 // 4 + self.hidden1 // 2 + self.hidden1, self.hidden1)
        self.GCN_GAT_bn = BatchNorm1d(self.hidden1)
        self.GCN_GAT_L = torch.nn.Linear(self.hidden1, 256)
        self.GCN_GAT_L_bn = BatchNorm1d(256)
        #最终拼接的线性层和结果层
        self.ALL_Linear1 = torch.nn.Linear(256, 128)
        self.ALL_Linear_bn1 = BatchNorm1d(128)
        self.ALL_Linear2 = torch.nn.Linear(128, 64)
        self.ALL_Linear_bn2 = BatchNorm1d(64)
        self.ALL_Linear3 = torch.nn.Linear(64, 32)
        self.ALL_Linear_bn3 = BatchNorm1d(32)
        self.ALL_Output = torch.nn.Linear(32, 1)
    def forward(self, data):
        x, edge_index, batch = data.x.to(device), data.edge_index.to(device), data.batch.to(device)
        #三个通道分别的传播路径
        #左 x_GCN_L_1
        x_GCN_L_1 = self.GCN_L_1(x, edge_index)
        x_GCN_L_1 = self.GCN_L_1_bn(x_GCN_L_1)
        x_GCN_L_1 = F.relu(x_GCN_L_1)
        x_GCN_L_1 = F.dropout(x_GCN_L_1, p = self.dropout_rate1, training=self.training)
        x_GCN_L_1 = F.dropout(F.relu(self.GAT_L_1_bn(self.GAT_L_1(x_GCN_L_1, edge_index))))
        #中 x_GCN_M_2
        x_GCN_M_1 = self.GCN_M_1(x, edge_index)
        x_GCN_M_1 = self.GCN_M_1_bn(x_GCN_M_1)
        x_GCN_M_1 = F.relu(x_GCN_M_1)
        x_GCN_M_1 = F.dropout(x_GCN_M_1, p= self.dropout_rate2, training=self.training)
        x_GCN_M_2 = self.GCN_M_2(x_GCN_M_1, edge_index)
        x_GCN_M_2 = self.GCN_M_2_bn(x_GCN_M_2)
        x_GCN_M_2 = F.relu(x_GCN_M_2)
        x_GCN_M_2 = F.dropout(x_GCN_M_2, p= self.dropout_rate3, training=self.training)
        x_GCN_M_2 = F.relu(self.GAT_M_2_bn(self.GAT_M_2(x_GCN_M_2, edge_index)))

        #右 x_GCN_R_3
        x_GCN_R_1 = self.GCN_R_1(x, edge_index)
        x_GCN_R_1 = self.GCN_R_1_bn(x_GCN_R_1)
        x_GCN_R_1 = F.relu(x_GCN_R_1)
        x_GCN_R_1 = F.dropout(x_GCN_R_1, p= self.dropout_rate4, training=self.training)
        x_GCN_R_2 = self.GCN_R_2(x_GCN_R_1, edge_index)
        x_GCN_R_2 = self.GCN_R_2_bn(x_GCN_R_2)
        x_GCN_R_2 = F.relu(x_GCN_R_2)
        x_GCN_R_2 = F.dropout(x_GCN_R_2, p= self.dropout_rate5, training=self.training)
        x_GCN_R_3 = self.GCN_R_3(x_GCN_R_2, edge_index)
        x_GCN_R_3 = self.GCN_R_3_bn(x_GCN_R_3)
        x_GCN_R_3 = F.relu(x_GCN_R_3)
        x_GCN_R_3 = F.relu(self.GAT_R_3_bn(self.GAT_R_3(x_GCN_R_3, edge_index)))
        #直接拼接
        x_input = torch.cat((x_GCN_L_1, x_GCN_M_2, x_GCN_R_3), dim = 1)
        x_ALL_GCN = self.GCN_GAT(x_input, edge_index)
        x_ALL_GCN = self.GCN_GAT_bn(x_ALL_GCN)
        x_ALL_GCN = F.relu(x_ALL_GCN)
        x_ALL_GCN = global_mean_pool(x_ALL_GCN, batch)
        x_ALL_GCN = self.GCN_GAT_L(x_ALL_GCN)
        x_ALL_GCN = self.GCN_GAT_L_bn(x_ALL_GCN)
        x_ALL_GCN = F.relu(x_ALL_GCN)
        x_ALL_GCN = F.dropout(x_ALL_GCN, p= self.dropout_rate6, training=self.training)
        x_linear1 = F.relu(self.ALL_Linear_bn1(self.ALL_Linear1(x_ALL_GCN)))
        x_linear1 = F.dropout(x_linear1, p= self.dropout_rate7, training=self.training)
        x_linear1 = F.relu(self.ALL_Linear_bn2(self.ALL_Linear2(x_linear1)))
        x_linear1 = F.dropout(x_linear1, p= self.dropout_rate8, training=self.training)
        x_linear1 = F.relu(self.ALL_Linear_bn3(self.ALL_Linear3(x_linear1)))
        x_linear1 = F.dropout(x_linear1, p= self.dropout_rate9, training=self.training)
        x_ALL_CAT = self.ALL_Output(x_linear1)
        return x_ALL_CAT
        
"""
class GCN(torch.nn.Module):
    def __init__(self, num_features, num_classes, init_method='xavier'):
        super(GCN, self).__init__()
        self.conv1 = GCNConv(num_features, 16)
        self.conv2 = GCNConv(16, num_classes)
        self.init_method = init_method
        self.initialize_weights()
    def initialize_weights(self):
        for m in self.modules():
            if isinstance(m, GCNConv):
                if self.init_method == 'xavier':
                    init.xavier_uniform_(m.weight)
                elif self.init_method == 'kaiming':
                    init.kaiming_uniform_(m.weight, nonlinearity='relu')

    def forward(self, x, edge_index):
        x = F.relu(self.conv1(x, edge_index))
        x = F.dropout(x, training=self.training)
        x = self.conv2(x, edge_index)
        return F.log_softmax(x, dim=1)
from torch_geometric.data import Data
import numpy as np
def create_random_graph(num_nodes, num_features, num_classes):
    edge_index = torch.randint(0, num_nodes, (2, num_nodes*2))
    x = torch.randn((num_nodes, num_features))
    y = torch.randint(0, num_classes, (num_nodes,))
    return Data(x=x, edge_index=edge_index, y=y)
def train_and_print_statistics(data, init_method):
    model = GCN(num_features=3, num_classes=2, init_method=init_method)
    optimizer = torch.optim.Adam(model.parameters(), lr=0.01)
    model.train()
    for epoch in range(10):
        optimizer.zero_grad()
        out = model(data.x, data.edge_index)
        loss = F.nll_loss(out[data.train_mask], data.y[data.train_mask])
        loss.backward()
        optimizer.step()
        print(f"Epoch {epoch+1} with {init_method} initialization:")
        for name, param in model.named_parameters():
            print(f"{name} - Mean: {param.data.mean()}, Std: {param.data.std()}")
data = create_random_graph(50, 3, 2)
data.train_mask = torch.rand(data.y.size(0)) < 0.8  # 随机训练掩码
train_and_print_statistics(data, 'xavier')
train_and_print_statistics(data, 'kaiming')

train_loader = DataLoader(graph_data_list, batch_size=batch_size_, sampler=train_subsampler)
for data in train_loader:
class GCNConv(MessagePassing):
    def __init__(self, in_channels, out_channels):
        super(GCNConv, self).__init__(aggr='add')  # "Add" aggregation.
        self.lin = torch.nn.Linear(in_channels, out_channels)
    def forward(self, x, edge_index):
        # x has shape [N, in_channels]
        # edge_index has shape [2, E]
        # Step 1: Add self-loops to the adjacency matrix.
        edge_index, _ = add_self_loops(edge_index, num_nodes=x.size(0))
        # Step 2: Linearly transform node feature matrix.
        x = self.lin(x)
        # Step 3: Compute normalization
        row, col = edge_index
        deg = degree(row, x.size(0), dtype=x.dtype)
        deg_inv_sqrt = deg.pow(-0.5)
        norm = deg_inv_sqrt[row] * deg_inv_sqrt[col]
        # Step 4-6: Start propagating messages.
        return self.propagate(edge_index, size=(x.size(0), x.size(0)), x=x,
                              norm=norm)
    def message(self, x_j, norm):
        # x_j has shape [E, out_channels]

        # Step 4: Normalize node features.
        return norm.view(-1, 1) * x_j

    def update(self, aggr_out):
        # aggr_out has shape [N, out_channels]

        # Step 6: Return new node embeddings.
        return aggr_out

conv = GCNConv(16, 32)

x = conv(x, edge_index)

import torch
import torch.nn.init as init
import torch.nn as nn
def NPI_with_KaimingHe(tensor, path):
    data_list = torch.load(path)
    vector_size = data_list[0].x.numel()
    EH = torch.zeros(vector_size)
    VH = torch.zeros(vector_size)
    VW = torch.zeros(vector_size)
    for i in data_list:
        flat_x = i.x.view(-1)
        EH += flat_x / len(data_list)
        VH += flat_x**2 / len(data_list)
        VW += flat_x / len(data_list)
    VH = VH - EH**2
    VW = (VW - EH**2) * (1 / 160)
    nn.init.kaiming_uniform_(tensor, a=0, mode='fan_in', nonlinearity='leaky_relu')
    adjustment = EH * VH / (VW**0.5)
    tensor.add_(adjustment.view_as(tensor))

class GCN(torch.nn.Module):
    def __init__(self, num_features, num_classes, init_method='kaiming'):
        super(GCN, self).__init__()
        self.conv = GCNConv(num_features, num_classes, init_method)
    def initialize_weights(self):
        for m in self.modules():
            if isinstance(m, GCNConv):
                if self.init_method == 'xavier':
                    init.xavier_uniform_(m.weight, nonlinearity='tanh')
                elif self.init_method == 'kaiming':
                    init.kaiming_uniform_(m.weight, nonlinearity='relu')
                elif self.init_method == 'NPI':
                    NPI_with_KaimingHe(model.conv.weight)
    def forward(self, data):
        x, edge_index = data.x, data.edge_index
        x = F.tanh(self.conv(x, edge_index))
        return F.log_softmax(x, dim=1)
def create_random_graph(num_nodes, num_features, num_classes):
    edge_index = torch.randint(0, num_nodes, (2, num_nodes * 2))  # Random edges
    x = torch.randn(num_nodes, num_features)  # Random features
    y = torch.randint(0, num_classes, (num_nodes,))  # Random labels
    return Data(x=x, edge_index=edge_index, y=y)
num_nodes = 100
num_features = 3
num_classes = 2
data = create_random_graph(num_nodes, num_features, num_classes)
model = GCN(num_features, num_classes,'NPI')
optimizer = optim.Adam(model.parameters(), lr=0.01)
criterion = torch.nn.CrossEntropyLoss()
model.train()
for epoch in range(10):
    optimizer.zero_grad()
    out = model(data)
    loss = criterion(out, data.y)
    loss.backward()
    optimizer.step()
    print(f"Epoch {epoch+1} weights statistics:")
    for name, param in model.named_parameters():
        print(f"{name} - Mean: {param.data.mean().item()}, Std: {param.data.std().item()}")
"""
from Bio.PDB import PDBParser
import networkx as nx
import os
import torch
from torch_geometric.data import Data
def find_pdb_files(dataset_path):
    labels = ["positive samples", "negative samples"]
    result = []
    for label in labels:
        label_path = os.path.join(dataset_path, label)
        for subdir, _, files in os.walk(label_path):
            subdir_name = os.path.basename(subdir)
            number = subdir_name.split('_')[0]
            for file in files:
                if 'rank_001' in file and file.endswith('.pdb'):
                    pdb_file_path = os.path.join(subdir, file)
                    result.append([pdb_file_path, number, label])
    return result
def pdb_to_connected_graph(pdb_file_path):
    parser = PDBParser()
    structure_id = os.path.basename(pdb_file_path).split('.')[0]
    structure = parser.get_structure(structure_id, pdb_file_path)
    model = structure[0]
    chain = next(model.get_chains())
    G = nx.Graph()
    residues = [residue for residue in chain if residue.get_resname() != 'HOH']
    num_residues = len(residues)
    for i in range(num_residues):
        for j in range(i + 1, num_residues):
            try:
                distance = residues[i]['CA'] - residues[j]['CA']
            except KeyError:
                continue
            if distance < 4.5:
                G.add_edge(i, j)
    is_connected = nx.is_connected(G)
    return G, is_connected

dataset_path = "C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\sample data"
pdb_files_info = find_pdb_files(dataset_path)
connectedlist = []
numberlist = []
graph_data_list = []  # Initialize the list to store PyG Data objects

for pdb_info in pdb_files_info:
    pdb_file_path, number, label = pdb_info
    if "positive samples" in pdb_file_path.split('\\'):
        print("There are positive samples here！")
        label__ = 1
    else:
        print("There are negative samples here")
        label__ = 0
    print("id:",pdb_file_path.split('\\')[-2].split("_")[0])
    aa = pdb_file_path.split('\\')[-2].split('_')[0]
    if label__ == 1:
        file_name = f'>ACP_{aa}|{label__}'
    elif label__ == 0:
        file_name = f'>non-ACP_{aa}|{label__}'
    # Assign a numerical label (1 for positive, 0 for negative)
    if label == 'positive samples':
        numeric_label = 1
    elif label == 'negative samples':
        numeric_label = 0
    else:
        raise ValueError(f"Unexpected label: {label}")
    print(f"{label}: {numeric_label}")  # This will print the label and its numeric representation
    graph, connected = pdb_to_connected_graph(pdb_file_path)
    print(graph)
    print(f"File Path: {pdb_file_path}")
    print(f"Number: {number}")
    print(f"Label: {label}")
    print(f"Is the protein structure graph connected? {connected}")
    connectedlist.append(connected)
    numberlist.append((label, number))
    num_nodes = len(graph.nodes)
    x = torch.ones((num_nodes, 1))  # Create a feature matrix with all ones
    y = torch.tensor([numeric_label], dtype=torch.long)  # Create a label tensor
    print(f"Data object label: {y.item()}")  # Print the label of the data object
    edge_index = torch.tensor(list(graph.edges), dtype=torch.long).t().contiguous()
    # Create a PyG Data object including the label
    data = Data(x=x, edge_index=edge_index, y=y,file_name = file_name)
    graph_data_list.append(data)
print(len(connectedlist))
print(False in connectedlist)
positive = []
negative = []
for i in numberlist:
    if i[0] == "positive samples":
        positive.append(int(i[1]))
    else:
        negative.append(int(i[1]))
for i in range(0, 380):
    if i not in positive:
        print("positive", i)
    if i not in negative:
        print("negative", i)
for i in graph_data_list:
    print(i.y)
torch.save(graph_data_list, 'graph_data_list.pt')

import torch
import matplotlib.pyplot as plt
import networkx as nx
from torch_geometric.utils import to_networkx
data_list = torch.load('graph_data_list.pt')
for i in data_list:
    print(i)
random_data = data_list[torch.randint(len(data_list), (1,)).item()]
print("Number of nodes:", random_data.num_nodes)
print("Number of edges:", random_data.num_edges)
print("Node features shape:", random_data.x.shape)
print("Edge indices shape:", random_data.edge_index.shape)
G = to_networkx(random_data, to_undirected=True)
# Draw the graph using matplotlib
plt.figure(figsize=(8, 6))
nx.draw(G, with_labels=True, node_color='lightblue', edge_color='gray')
plt.title('Randomly Selected Graph from Data List')
plt.show()

import networkx as nx
from networkx.algorithms import cluster, centrality
from karateclub import DeepWalk
import numpy as np
import torch
from torch_geometric.data import Data
from torch_geometric.utils import to_networkx
def normalize_rows(tensor):
    """
    Normalize the rows of a 2D PyTorch Tensor.
    Each row is normalized to have values between 0 and 1.
    Parameters:
    tensor (torch.Tensor): The 2D Tensor to normalize.
    Returns:
    torch.Tensor: The row-normalized 2D Tensor.
    """
    # Ensure the tensor is a floating point type for division
    tensor = tensor.float()
    # Calculate the min and max values along each row
    min_vals = torch.min(tensor, dim=1, keepdim=True).values
    max_vals = torch.max(tensor, dim=1, keepdim=True).values
    # Perform row-wise normalization
    normalized_tensor = (tensor - min_vals) / (max_vals - min_vals)
    return normalized_tensor
def normalize_rows_l2(tensor):
    """
    Normalize the rows of a 2D PyTorch Tensor using L2 norm.
    Each row is divided by its L2 norm, making the norm of the row vector equal to 1.
    Parameters:
    tensor (torch.Tensor): The 2D Tensor to normalize.
    Returns:
    torch.Tensor: The row-normalized 2D Tensor.
    """
    # Calculate the L2 norm for each row (add a small epsilon for numerical stability)
    l2_norms = torch.norm(tensor, p=2, dim=1, keepdim=True) + 1e-10
    # Divide each row by its L2 norm
    normalized_tensor = tensor / l2_norms
    return normalized_tensor
def normalize_rows_z_score(tensor):
    """
    Normalize the rows of a 2D PyTorch Tensor using Z-Score normalization.
    Each row will have a mean of 0 and a standard deviation of 1.
    Parameters:
    tensor (torch.Tensor): The 2D Tensor to normalize.
    Returns:
    torch.Tensor: The row-normalized 2D Tensor.
    """
    # Ensure the tensor is a floating point type for division
    tensor = tensor.float()
    # Calculate the mean and standard deviation for each row
    mean = torch.mean(tensor, dim=1, keepdim=True)
    std = torch.std(tensor, dim=1, keepdim=True) + 1e-10  # Avoid division by zero
    # Perform row-wise Z-Score normalization
    normalized_tensor = (tensor - mean) / std
    return normalized_tensor
def normalize_by_global_max(tensor):
    """
    Normalize a 2D PyTorch Tensor by the global maximum value.
    All elements in the tensor are normalized to have values between 0 and 1 based on the global maximum value.
    Parameters:
    tensor (torch.Tensor): The 2D Tensor to normalize.
    Returns:
    torch.Tensor: The globally normalized 2D Tensor.
    """
    # Ensure the tensor is a floating point type for division
    tensor = tensor.float()
    # Calculate the global max value
    global_max = torch.max(tensor)
    # Perform global normalization
    normalized_tensor = tensor / global_max
    return normalized_tensor

Graph_Features_List = []
model = DeepWalk(dimensions=80)
Graph_list = torch.load("C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\graph_data_list.pt")
for data in Graph_list:
    G = to_networkx(data, to_undirected=True)
    eigenvector_centrality = torch.tensor(list(nx.eigenvector_centrality(G, max_iter=100000).values()),
                                          dtype=torch.float32).view(G.number_of_nodes(), -1)
    betweenness_centrality = torch.tensor(list(nx.betweenness_centrality(G).values()), dtype=torch.float32).view(
        G.number_of_nodes(), -1)
    closeness_centrality = torch.tensor(list(nx.closeness_centrality(G).values()), dtype=torch.float32).view(
        G.number_of_nodes(), -1)
    clustering_coefficient = torch.tensor(list(nx.clustering(G).values()), dtype=torch.float32).view(
        G.number_of_nodes(), -1)
    adj_matrix = nx.normalized_laplacian_matrix(G)
    eigenvalues, eigenvectors = np.linalg.eigh(adj_matrix.toarray())
    k = 5
    feature_vectors_5 = torch.tensor(eigenvectors[:, :k], dtype=torch.float32).view(G.number_of_nodes(), -1)
    pagerank = torch.tensor(list(nx.pagerank(G).values()), dtype=torch.float32).view(G.number_of_nodes(), -1)
    k_core = torch.tensor(list(nx.core_number(G).values()), dtype=torch.float32).view(G.number_of_nodes(),-1)
    k_core_normalize_rows_z_score = normalize_by_global_max(k_core)
    degrees = torch.tensor(list(dict(G.degree()).values()), dtype=torch.float32).view(G.number_of_nodes(), -1)
    degrees_normalize_rows_z_score = normalize_by_global_max(degrees)
    degree_centrality = torch.tensor(list(centrality.degree_centrality(G).values()), dtype=torch.float32).view(
        G.number_of_nodes(), -1)
    clustering_coefficient = torch.tensor(list(cluster.clustering(G).values()), dtype=torch.float32).view(
        G.number_of_nodes(), -1)
    model = DeepWalk(dimensions=86)
    model.fit(G)
    embedding = torch.tensor(model.get_embedding(), dtype=torch.float32).view(G.number_of_nodes(), -1)
    print(betweenness_centrality.shape)
    print(closeness_centrality.shape)
    print(clustering_coefficient.shape)
    print(feature_vectors_5.shape)
    print(pagerank.shape)
    print(k_core_normalize_rows_z_score.shape)
    print(degrees_normalize_rows_z_score.shape)
    print(degree_centrality.shape)
    print(embedding.shape)
    Graph_Features = torch.cat((data.x,betweenness_centrality,closeness_centrality,feature_vectors_5,pagerank,k_core_normalize_rows_z_score,degrees_normalize_rows_z_score,degree_centrality,clustering_coefficient,embedding),dim=1)
    Graph_Features_List.append(Graph_Features)
    data.x = Graph_Features

print(len(Graph_Features_List))
print(Graph_Features_List[0].shape)
print(Graph_Features_List[0])
torch.save(Graph_list,"C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\graph_data_list.pt")

import torch
import matplotlib.pyplot as plt
import networkx as nx
from torch_geometric.utils import to_networkx
from torch_geometric.data import Data
# Load the list of PyG Data objects from the .pt file
data_list = torch.load('C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\graph_data_list.pt')
sequences = {}
with open("C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\Benchmark dataset\\Benchmark dataset.txt", 'r') as f:
    content = f.read().splitlines()
# Assuming that each header is unique and directly followed by its sequence
for i in range(0, len(content), 2):
    header = content[i].strip()
    seq = content[i + 1].strip()
    sequences[header] = seq
    #print(header)
    #print(seq)
    #print(sequences[header])
print(data_list)
for i in data_list:
    try:
        i.seq = sequences[i.file_name]
    except KeyError:
        print(i.file_name)
for i in data_list:
    print(i.file_name)
data_list = [data for data in data_list if data.file_name != "'>ACP_175|1'"]
print(data_list)
for i in data_list:
    print(i.file_name)
for i in data_list:
    print(i)
torch.save(data_list, 'C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\graph_data_list.pt')

import torch
import numpy as np

def sequence_to_tensor(sequence, file_path):
    # Load the AAindex data
    aaindex_dict = {}
    with open(file_path, 'r') as file:
        next(file)  # Skip the title line
        for line in file:
            parts = line.split()
            if parts:  # Ensure the line is not empty
                # Convert AAindex values to float32 and store in the dictionary
                aaindex_dict[parts[0]] = np.array(parts[1:], dtype=np.float32)
    # Map the sequence to a matrix using the AAindex data
    matrix = []
    for char in sequence:
        if char in aaindex_dict:
            matrix.append(aaindex_dict[char])
        else:
            print(f"Warning: '{char}' not found in AAindex. Skipping.")
    # Convert the list of numpy arrays to a single numpy array
    matrix_np = np.stack(matrix)
    # Convert the numpy array to a PyTorch tensor of dtype float32
    tensor = torch.tensor(matrix_np, dtype=torch.float32)
    return tensor
blosum62 = {
            'A': [4, -1, -2, -2, 0, -1, -1, 0, -2, -1, -1, -1, -1, -2, -1, 1, 0, -3, -2, 0],  # A
            'R': [-1, 5, 0, -2, -3, 1, 0, -2, 0, -3, -2, 2, -1, -3, -2, -1, -1, -3, -2, -3],  # R
            'N': [-2, 0, 6, 1, -3, 0, 0, 0, 1, -3, -3, 0, -2, -3, -2, 1, 0, -4, -2, -3],  # N
            'D': [-2, -2, 1, 6, -3, 0, 2, -1, -1, -3, -4, -1, -3, -3, -1, 0, -1, -4, -3, -3],  # D
            'C': [0, -3, -3, -3, 9, -3, -4, -3, -3, -1, -1, -3, -1, -2, -3, -1, -1, -2, -2, -1],  # C
            'Q': [-1, 1, 0, 0, -3, 5, 2, -2, 0, -3, -2, 1, 0, -3, -1, 0, -1, -2, -1, -2],  # Q
            'E': [-1, 0, 0, 2, -4, 2, 5, -2, 0, -3, -3, 1, -2, -3, -1, 0, -1, -3, -2, -2],  # E
            'G': [0, -2, 0, -1, -3, -2, -2, 6, -2, -4, -4, -2, -3, -3, -2, 0, -2, -2, -3, -3],  # G
            'H': [-2, 0, 1, -1, -3, 0, 0, -2, 8, -3, -3, -1, -2, -1, -2, -1, -2, -2, 2, -3],  # H
            'I': [-1, -3, -3, -3, -1, -3, -3, -4, -3, 4, 2, -3, 1, 0, -3, -2, -1, -3, -1, 3],  # I
            'L': [-1, -2, -3, -4, -1, -2, -3, -4, -3, 2, 4, -2, 2, 0, -3, -2, -1, -2, -1, 1],  # L
            'K': [-1, 2, 0, -1, -3, 1, 1, -2, -1, -3, -2, 5, -1, -3, -1, 0, -1, -3, -2, -2],  # K
            'M': [-1, -1, -2, -3, -1, 0, -2, -3, -2, 1, 2, -1, 5, 0, -2, -1, -1, -1, -1, 1],  # M
            'F': [-2, -3, -3, -3, -2, -3, -3, -3, -1, 0, 0, -3, 0, 6, -4, -2, -2, 1, 3, -1],  # F
            'P': [-1, -2, -2, -1, -3, -1, -1, -2, -2, -3, -3, -1, -2, -4, 7, -1, -1, -4, -3, -2],  # P
            'S': [1, -1, 1, 0, -1, 0, 0, 0, -1, -2, -2, 0, -1, -2, -1, 4, 1, -3, -2, -2],  # S
            'T': [0, -1, 0, -1, -1, -1, -1, -2, -2, -1, -1, -1, -1, -2, -1, 1, 5, -2, -2, 0],  # T
            'W': [-3, -3, -4, -4, -2, -2, -3, -2, -2, -3, -2, -3, -1, 1, -4, -3, -2, 11, 2, -3],  # W
            'Y': [-2, -2, -2, -3, -2, -1, -2, -3, 2, -1, -1, -2, -1, 3, -3, -2, -2, 2, 7, -1],  # Y
            'V': [0, -3, -3, -3, -1, -2, -2, -3, -3, 3, 1, -2, 1, -1, -2, -2, 0, -3, -1, 4],  # V
            'O': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],  # O
        }
SD_dict = {
            'A': [1,0,0,0,0,0,0],  # A
            'R': [0,0,0,0,1,0,0],  # R
            'N': [0,0,0,1,0,0,0],  # N
            'D': [0,0,0,0,0,1,0],  # D
            'C': [0,0,0,0,0,0,1],  # C
            'Q': [0,0,0,1,0,0,0],  # Q
            'E': [0,0,0,0,0,1,0],  # E
            'G': [1,0,0,0,0,0,0],  # G
            'H': [0,0,0,1,0,0,0],  # H
            'I': [0,1,0,0,0,0,0],  # I
            'L': [0,1,0,0,0,0,0],  # L
            'K': [0,0,0,0,1,0,0],  # K
            'M': [0,0,1,0,0,0,0],  # M
            'F': [0,1,0,0,0,0,0],  # F
            'P': [0,1,0,0,0,0,0],  # P
            'S': [0,0,1,0,0,0,0],  # S
            'T': [0,0,1,0,0,0,0],  # T
            'W': [0,0,0,1,0,0,0],  # W
            'Y': [0,0,1,0,0,0,0],  # Y
            'V': [1,0,0,0,0,0,0],  # V
            'O': [0,0,0,0,0,0,0],  # O
        }
PC_dict = {
            'A': [1, 0, 0, 0, 0],  # A
            'R': [0, 0, 1, 0, 0],  # R
            'N': [0, 0, 0, 0, 1],  # N
            'D': [0, 0, 0, 1, 0],  # D
            'C': [0, 0, 0, 0, 1],  # C
            'Q': [0, 0, 0, 0, 1],  # Q
            'E': [0, 0, 0, 1, 0],  # E
            'G': [1, 0, 0, 0, 0],  # G
            'H': [0, 0, 1, 0, 0],  # H
            'I': [1, 0, 0, 0, 0],  # I
            'L': [1, 0, 0, 0, 0],  # L
            'K': [0, 0, 1, 0, 0],  # K
            'M': [1, 0, 0, 0, 0],  # M
            'F': [0, 1, 0, 0, 0],  # F
            'P': [0, 0, 0, 0, 1],  # P
            'S': [0, 0, 0, 0, 1],  # S
            'T': [0, 0, 0, 0, 1],  # T
            'W': [0, 1, 0, 0, 0],  # W
            'Y': [0, 1, 0, 0, 0],  # Y
            'V': [1, 0, 0, 0, 0],  # V
            'O': [0, 0, 0, 0, 0],  # O
        }
EGB_dict = {
            'A': [1, 1, 1],  # A
            'R': [0, 0, 1],  # R
            'N': [1, 0, 0],  # N
            'D': [0, 1, 0],  # D
            'C': [1, 0, 0],  # C
            'Q': [1, 0, 0],  # Q
            'E': [0, 1, 0],  # E
            'G': [1, 1, 1],  # G
            'H': [0, 0, 1],  # H
            'I': [1, 1, 1],  # I
            'L': [1, 1, 1],  # L
            'K': [0, 0, 1],  # K
            'M': [1, 1, 1],  # M
            'F': [1, 1, 1],  # F
            'P': [1, 1, 1],  # P
            'S': [1, 0, 0],  # S
            'T': [1, 0, 0],  # T
            'W': [1, 1, 1],  # W
            'Y': [1, 0, 0],  # Y
            'V': [1, 1, 1],  # V
            'O': [0, 0, 0],  # O
        }
PAM250 = {
            'A': [2,-2,0,0,-3,1,-1,-1,-1,-2,-1,0,1,0,-2,1,1,0,-6,-3],  #A
            'R': [-2,-4,-1,-1,-4,-3,2,-2,3,-3,0,0,0,1,6,0,-1,-2,2,-4],  # R
            'N': [0,-4,2,1,-3,0,2,-2,1,-3,-2,2,0,1,0,1,0,-2,-4,-2],  # N
            'D': [0,-5,4,3,-6,1,1,-2,0,-4,-3,2,-1,2,-1,0,0,-2,-7,-4],  # D
            'C': [-2,12,-5,-5,-4,-3,-3,-2,-5,-6,-5,-4,-3,-5,-4,0,-2,-2,-8,0],  # C
            'Q': [0,-5,2,2,-5,-1,3,-2,1,-2,-1,1,0,4,1,-1,-1,-2,-5,-4],  # Q
            'E': [0,-5,3,4,-5,0,1,-2,0,-3,-2,1,-1,2,-1,0,0,-2,-7,-4],  # E
            'G': [1,-3,1,0,-5,5,-2,-3,-2,-4,-3,0,0,-1,-3,1,0,-1,-7,-5],  # G
            'H': [-1,-3,1,1,-2,-2,6,-2,0,-2,-2,2,0,3,2,-1,-1,-2,-3,0],  # H
            'I': [-1,-2,-2,-2,1,-3,-2,5,-2,2,2,-2,-2,-2,-2,-1,0,4,-5,-1],  # I
            'L': [-2,-6,-4,-3,2,-4,-2,2,-3,6,4,-3,-3,-2,-3,-3,-2,2,-2,-1],  # L
            'K': [-1,-5,0,0,-5,-2,0,-2,5,-3,0,1,-1,1,3,0,0,-2,-3,-4],  # K
            'M': [-1,-5,-3,-2,0,-3,-2,2,0,4,6,-2,-2,-1,0,-2,-1,2,-4,-2],  # M
            'F': [-3,-4,-6,-5,9,-5,-2,1,-5,2,0,-3,-5,-5,-4,-3,-3,-1,0,7],  # F
            'P': [1,-3,-1,-1,-5,0,0,-2,-1,-3,-2,0,6,0,0,1,0,-1,-6,-5],  # P
            'S': [1,0,0,0,-3,1,-1,-1,0,-3,-2,1,1,-1,0,2,1,-1,-2,-3],  # S
            'T': [1,-2,0,0,-3,0,-1,0,0,-2,-1,0,0,-1,-1,1,3,0,-5,-3],  # T
            'W': [-6,-8,-7,-7,0,-7,-3,-5,-3,-2,-4,-4,-6,-5,2,-2,-5,-6,17,0],  # W
            'Y': [-3,0,-4,-4,7,-5,0,-1,-4,-1,-2,-2,-5,-4,-4,-3,-3,-2,0,10],  # Y
            'V': [0,-2,-2,-2,-1,-1,-2,4,-2,2,2,-2,-1,-2,-2,-1,0,4,-6,-2],  # V
            'O': [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],  # O
        }
zscale = {
            'A': [0.24, -2.32, 0.60, -0.14, 1.30],  # A
            'C': [0.84, -1.67, 3.71, 0.18, -2.65],  # C
            'D': [3.98, 0.93, 1.93, -2.46, 0.75],  # D
            'E': [3.11, 0.26, -0.11, -0.34, -0.25],  # E
            'F': [-4.22, 1.94, 1.06, 0.54, -0.62],  # F
            'G': [2.05, -4.06, 0.36, -0.82, -0.38],  # G
            'H': [2.47, 1.95, 0.26, 3.90, 0.09],  # H
            'I': [-3.89, -1.73, -1.71, -0.84, 0.26],  # I
            'K': [2.29, 0.89, -2.49, 1.49, 0.31],  # K
            'L': [-4.28, -1.30, -1.49, -0.72, 0.84],  # L
            'M': [-2.85, -0.22, 0.47, 1.94, -0.98],  # M
            'N': [3.05, 1.62, 1.04, -1.15, 1.61],  # N
            'P': [-1.66, 0.27, 1.84, 0.70, 2.00],  # P
            'Q': [1.75, 0.50, -1.44, -1.34, 0.66],  # Q
            'R': [3.52, 2.50, -3.50, 1.99, -0.17],  # R
            'S': [2.39, -1.07, 1.15, -1.39, 0.67],  # S
            'T': [0.75, -2.18, -1.12, -1.46, -0.40],  # T
            'V': [-2.59, -2.64, -1.54, -0.85, -0.02],  # V
            'W': [-4.36, 3.94, 0.59, 3.44, -1.59],  # W
            'Y': [-2.54, 2.44, 0.43, 0.04, -1.47],  # Y
            'O': [0.00, 0.00, 0.00, 0.00, 0.00],  # -
        }
def encode_sequence(sequence, EncodeWays):
    EncodeWays_ = EncodeWays
    matrix = []
    for char in sequence:
        if char in EncodeWays_:
            matrix.append(EncodeWays_[char])
        else:
            print(f"Warning: '{char}' not found in EGB encoding. Using [0, 0, 0] as placeholder.")
            matrix.append([0, 0, 0])  # Using a placeholder for unknown characters.
    return torch.tensor(matrix, dtype=torch.float32)
data_list = torch.load('C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\graph_data_list.pt')
from gensim.models import Word2Vec
import numpy as np
sequences = [data.seq for data in data_list]
model = Word2Vec(sequences, vector_size=50, window=5, min_count=1, workers=4)
def encode_with_word2vec(sequence, model):
    encoded_sequence = np.vstack([model.wv[word] if word in model.wv else np.zeros((model.vector_size,)) for word in sequence])
    return torch.tensor(encoded_sequence, dtype=torch.float32)
def encode_and_concatenate_sequences(sequence, encoding_dictionaries, word2vec_model):
    encoded_parts = []
    for dictionary in encoding_dictionaries:
        encoded_sequence = np.array(
            [dictionary.get(aa, np.zeros_like(next(iter(dictionary.values())))) for aa in sequence])
        encoded_parts.append(torch.tensor(encoded_sequence, dtype=torch.float32))
    w2v_encoded = encode_with_word2vec(sequence, word2vec_model)
    final_encoding = torch.cat(encoded_parts + [w2v_encoded], dim=1)
    return final_encoding
for i in data_list:
    sequence = i.seq
    encoding_dictionaries = [blosum62, SD_dict, PC_dict, EGB_dict, PAM250, zscale]
    w2v_encoded = encode_with_word2vec(sequence, model)
    i.x = torch.cat((i.x, encode_and_concatenate_sequences(sequence, encoding_dictionaries, model),sequence_to_tensor(sequence, 'AAindex.txt')), dim=1)
    print(i)
print(data_list)
torch.save(data_list,'C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\graph_data_list.pt')
import os
import torch
import numpy as np
from torch_geometric.data import Data
from sklearn.ensemble import RandomForestClassifier
from torch_geometric.nn import global_mean_pool
def select_and_save_graph_features(data_path, max_features=None, step=16, save_folder='Select'):
    # Load the data
    graph_data_list = torch.load(data_path)
    # Prepare the lists for features and labels
    x_features = []
    y_labels = []
    # Extract features and labels
    for data in graph_data_list:
        batch = torch.zeros(data.x.size(0), dtype=torch.long)
        global_features = global_mean_pool(data.x, batch)
        x_features.append(global_features.numpy()[0])
        y_labels.append(data.y.numpy()[0])
    # Convert features to a numpy array
    x_features = np.array(x_features)
    # Initialize and fit the Random Forest classifier
    rf = RandomForestClassifier()
    rf.fit(x_features, y_labels)
    # Get the feature importances
    importances = rf.feature_importances_
    # If max_features is not set, use the total number of features
    if max_features is None:
        max_features = x_features.shape[1]
    # Create the select folder if it does not exist
    select_folder_path = os.path.join(os.path.dirname(data_path), save_folder)
    os.makedirs(select_folder_path, exist_ok=True)
    # Create a dictionary to store the new data lists
    new_data_lists = {}
    # Iterate over the number of features to select
    for num_features in range(step, max_features + 1, step):
        # Get the indices of the top features
        indices = np.argsort(importances)[::-1][:num_features]
        # Create a mask for the selected features
        mask = np.zeros_like(importances, dtype=bool)
        mask[indices] = True
        # Select the features using the mask
        X_selected = x_features[:, mask]
        # Create a new list for the new graph data with selected features
        new_data_list = []
        for data in graph_data_list:
            new_x = data.x[:, mask]
            new_data = Data(x=new_x, edge_index=data.edge_index, y=data.y, seq=data.seq)
            new_data_list.append(new_data)
        # Save the new data list to the select folder
        filename = f"graph_data_list_seq_onehot_GF_select_{num_features}.pt"
        save_path = os.path.join(select_folder_path, filename)
        torch.save(new_data_list, save_path)
        # Store the path in the dictionary
        new_data_lists[filename] = save_path
    return new_data_lists
# Example usage:
data_path = 'C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\graph_data_list.pt'
new_data_lists = select_and_save_graph_features(data_path)
from torch.utils.data import SubsetRandomSampler
from torch_geometric.loader import DataLoader
from sklearn.model_selection import KFold
from openpyxl import load_workbook
import pandas as pd
import openpyxl
from torch.optim import Adam
from sklearn.metrics import confusion_matrix, matthews_corrcoef, accuracy_score
import numpy as np
import matplotlib.pyplot as plt
from torch.nn.modules.batchnorm import _BatchNorm
import torch
import torch.nn.functional as F
from torch_geometric.nn import GCNConv, GATConv, global_mean_pool, DirGNNConv, PointGNNConv, AGNNConv ,GNNFF
from torch.nn import BatchNorm1d
excel_path = "C:\\Users\\ASUS\\Desktop\\MGNN-NPI\\Result.xlsx"

def calculate_average_mcc(mcc_results):
    num_epochs = len(mcc_results[0])
    mcc_averages = [0] * num_epochs
    for epoch in range(num_epochs):
        mcc_sum = sum(fold_mcc[epoch] for fold_mcc in mcc_results)
        mcc_averages[epoch] = mcc_sum / len(mcc_results)
    return mcc_averages
class NodeLevelBatchNorm(_BatchNorm):
    def __init__(self, num_features, eps=1e-5, momentum=0.1, affine=True,
                 track_running_stats=True):
        super(NodeLevelBatchNorm, self).__init__(
            num_features, eps, momentum, affine, track_running_stats)
    def _check_input_dim(self, input):
        if input.dim() != 2:
            raise ValueError('expected 2D input (got {}D input)'
                             .format(input.dim()))
    def forward(self, input):
        self._check_input_dim(input)
        if self.momentum is None:
            exponential_average_factor = 0.0
        else:
            exponential_average_factor = self.momentum
        if self.training and self.track_running_stats:
            if self.num_batches_tracked is not None:
                self.num_batches_tracked = self.num_batches_tracked + 1
                if self.momentum is None:
                    exponential_average_factor = 1.0 / float(self.num_batches_tracked)
                else:
                    exponential_average_factor = self.momentum
        return torch.functional.F.batch_norm(
            input, self.running_mean, self.running_var, self.weight, self.bias,
            self.training or not self.track_running_stats,
            exponential_average_factor, self.eps)
    def extra_repr(self):
        return 'num_features={num_features}, eps={eps}, ' \
               'affine={affine}'.format(**self.__dict__)
import os
class GCN(torch.nn.Module):
    def __init__(self, input_dim, dropout_rate):
        super(GCN, self).__init__()
        self.input_dim = input_dim
        self.dropout_rate = dropout_rate
        self.hidden1 = self.input_dim - self.input_dim % 16
        self.GCN_L_1 = GCNConv(self.input_dim, self.hidden1)
        self.GCN_L_1_bn = BatchNorm1d(self.hidden1)
        self.GAT_L_1 = GATConv(self.hidden1, self.hidden1)
        self.GAT_L_1_bn = BatchNorm1d(self.hidden1)
        self.GCN_M_1 = GCNConv(self.input_dim, self.hidden1)
        self.GCN_M_1_bn = BatchNorm1d(self.hidden1)
        self.GCN_M_2 = GCNConv(self.hidden1, self.hidden1 // 2)
        self.GCN_M_2_bn = BatchNorm1d(self.hidden1 // 2)
        self.GAT_M_2 = GATConv(self.hidden1 // 2, self.hidden1 // 2)
        self.GAT_M_2_bn = BatchNorm1d(self.hidden1 // 2)
        self.GCN_R_1 = GCNConv(self.input_dim, self.hidden1)
        self.GCN_R_1_bn = BatchNorm1d(self.hidden1)
        self.GCN_R_2 = GCNConv(self.hidden1, self.hidden1 // 2)
        self.GCN_R_2_bn = BatchNorm1d(self.hidden1 // 2)
        self.GCN_R_3 = GCNConv(self.hidden1 // 2, self.hidden1 // 4)
        self.GCN_R_3_bn = BatchNorm1d(self.hidden1 // 4)
        self.GAT_R_3 = GATConv(self.hidden1 // 4, self.hidden1 // 4)
        self.GAT_R_3_bn = BatchNorm1d(self.hidden1 // 4)
        self.GCN_GAT = GATConv(self.hidden1 // 4 + self.hidden1 // 2 + self.hidden1, self.hidden1)
        self.GCN_GAT_bn = BatchNorm1d(self.hidden1)
        self.GCN_GAT_L = torch.nn.Linear(self.hidden1, 256)
        self.GCN_GAT_L_bn = BatchNorm1d(256)
        self.ALL_Linear1 = torch.nn.Linear(256, 128)
        self.ALL_Linear_bn1 = BatchNorm1d(128)
        self.ALL_Linear2 = torch.nn.Linear(128, 64)
        self.ALL_Linear_bn2 = BatchNorm1d(64)
        self.ALL_Linear3 = torch.nn.Linear(64, 32)
        self.ALL_Linear_bn3 = BatchNorm1d(32)
        self.ALL_Output = torch.nn.Linear(32, 1)
    def forward(self, data):
        x, edge_index, batch = data.x.to(device), data.edge_index.to(device), data.batch.to(device)
        x_GCN_L_1 = self.GCN_L_1(x, edge_index)
        x_GCN_L_1 = self.GCN_L_1_bn(x_GCN_L_1)
        x_GCN_L_1 = F.tanh(x_GCN_L_1)
        x_GCN_L_1 = F.dropout(x_GCN_L_1, p = 0.25, training=self.training)
        x_GCN_L_1 = F.dropout(F.tanh(self.GAT_L_1_bn(self.GAT_L_1(x_GCN_L_1, edge_index))))
        x_GCN_M_1 = self.GCN_M_1(x, edge_index)
        x_GCN_M_1 = self.GCN_M_1_bn(x_GCN_M_1)
        x_GCN_M_1 = F.tanh(x_GCN_M_1)
        x_GCN_M_1 = F.dropout(x_GCN_M_1, p=0.25, training=self.training)
        x_GCN_M_2 = self.GCN_M_2(x_GCN_M_1, edge_index)
        x_GCN_M_2 = self.GCN_M_2_bn(x_GCN_M_2)
        x_GCN_M_2 = F.tanh(x_GCN_M_2)
        x_GCN_M_2 = F.dropout(x_GCN_M_2, p=0.2, training=self.training)
        x_GCN_M_2 = F.tanh(self.GAT_M_2_bn(self.GAT_M_2(x_GCN_M_2, edge_index)))
        x_GCN_R_1 = self.GCN_R_1(x, edge_index)
        x_GCN_R_1 = self.GCN_R_1_bn(x_GCN_R_1)
        x_GCN_R_1 = F.relu(x_GCN_R_1)
        x_GCN_R_1 = F.dropout(x_GCN_R_1, p=0.2, training=self.training)
        x_GCN_R_2 = self.GCN_R_2(x_GCN_R_1, edge_index)
        x_GCN_R_2 = self.GCN_R_2_bn(x_GCN_R_2)
        x_GCN_R_2 = F.relu(x_GCN_R_2)
        x_GCN_R_2 = F.dropout(x_GCN_R_2, p=0.2, training=self.training)
        x_GCN_R_3 = self.GCN_R_3(x_GCN_R_2, edge_index)
        x_GCN_R_3 = self.GCN_R_3_bn(x_GCN_R_3)
        x_GCN_R_3 = F.relu(x_GCN_R_3)
        x_GCN_R_3 = F.relu(self.GAT_R_3_bn(self.GAT_R_3(x_GCN_R_3, edge_index)))
        x_input = torch.cat((x_GCN_L_1, x_GCN_M_2, x_GCN_R_3), dim = 1)
        x_ALL_GCN = self.GCN_GAT(x_input, edge_index)
        x_ALL_GCN = self.GCN_GAT_bn(x_ALL_GCN)
        x_ALL_GCN = F.relu(x_ALL_GCN)
        x_ALL_GCN = global_mean_pool(x_ALL_GCN, batch)
        x_ALL_GCN = self.GCN_GAT_L(x_ALL_GCN)
        x_ALL_GCN = self.GCN_GAT_L_bn(x_ALL_GCN)
        x_ALL_GCN = F.relu(x_ALL_GCN)
        x_ALL_GCN = F.dropout(x_ALL_GCN, p=0.15, training=self.training)
        x_linear1 = F.relu(self.ALL_Linear_bn1(self.ALL_Linear1(x_ALL_GCN)))
        x_linear1 = F.dropout(x_linear1, p=0.15, training=self.training)
        x_linear1 = F.relu(self.ALL_Linear_bn2(self.ALL_Linear2(x_linear1)))
        x_linear1 = F.dropout(x_linear1, p=0.15, training=self.training)
        x_linear1 = F.relu(self.ALL_Linear_bn3(self.ALL_Linear3(x_linear1)))
        x_linear1 = F.dropout(x_linear1, p=0.15, training=self.training)
        x_ALL_CAT = self.ALL_Output(x_linear1)
        #x_ALL_CAT = torch.sigmoid(x_ALL_CAT)
        #x_pool_L = global_mean_pool(x_GCN_L_1, batch)
        #x_pool_M = global_mean_pool(x_GCN_M_2, batch)
        #x_pool_R = global_mean_pool(x_GCN_R_3, batch)
        #x_CAT_pool = torch.cat((x_pool_L, x_pool_M, x_pool_R), dim = 1)
        #x_GCN_GAT = torch.cat((x_GCN_L_1, x_GCN_M_2, x_GCN_R_3), dim=1)
        #x_CAT_pool = self.CAT_pool_L(x_CAT_pool)
        #x_CAT_pool = self.CAT_pool_L_bn(x_CAT_pool)
        #x_CAT_pool = F.relu(x_CAT_pool)
        #x_CAT_pool = F.dropout(x_CAT_pool, p=self.dropout_rate, training=self.training)
        #x_GCN_GAT = self.GCN_GAT(x_GCN_GAT, edge_index)
        #x_GCN_GAT = self.GCN_GAT_bn(x_GCN_GAT)
        #x_GCN_GAT = F.relu(x_GCN_GAT)
        #x_GCN_GAT_pool = global_mean_pool(x_GCN_GAT, batch)
        #x_GCN_GAT_pool = self.GCN_GAT_L(x_GCN_GAT_pool)
        #x_GCN_GAT_pool = self.GCN_GAT_L_bn(x_GCN_GAT_pool)
        #x_GCN_GAT_pool = F.relu(x_GCN_GAT_pool)
        #x_GCN_GAT_pool = F.dropout(x_GCN_GAT_pool, p=self.dropout_rate, training=self.training)
        #x_L_L = self.L_L(x_pool_L)
        #x_L_L = self.L_L_bn(x_L_L)
        #x_L_L = F.relu(x_L_L)
        #x_L_L = F.dropout(x_L_L, p=self.dropout_rate, training=self.training)
        #x_M_L = self.M_L(x_pool_M)
        #x_M_L = self.M_L_bn(x_M_L)
        #x_M_L = F.relu(x_M_L)
        #x_M_L = F.dropout(x_M_L, p = self.dropout_rate, training = self.training)
        #x_R_L = self.R_L(x_pool_R)
        #x_R_L = self.R_L_bn(x_R_L)
        #x_R_L = F.relu(x_R_L)
        #x_R_L = F.dropout(x_R_L, p = self.dropout_rate, training = self.training)
        #x_L_CAT = torch.cat((x_L_L, x_M_L, x_R_L), dim = 1)
        #_L_CAT = self.L_CAT_L(x_L_CAT)
        #_L_CAT = self.L_CAT_L_bn(x_L_CAT)
        #_L_CAT = F.relu(x_L_CAT)
        #_L_CAT = F.dropout(x_L_CAT, p = self.dropout_rate, training = self.training)
        #x_ALL_CAT = torch.cat((x_CAT_pool, x_GCN_GAT_pool, x_L_CAT), dim = 1)
        #x_ALL_CAT = self.ALL_Linear(x_ALL_CAT)
        #x_ALL_CAT = self.ALL_Linear_bn(x_ALL_CAT)
        #x_ALL_CAT = F.relu(x_ALL_CAT)
        #x_ALL_CAT = F.dropout(x_ALL_CAT)
        #x_ALL_CAT = self.ALL_Output(x_ALL_CAT)
        #x_ALL_CAT = torch.sigmoid(x_ALL_CAT)
        return x_ALL_CAT
folder_path = r'C:\Users\ASUS\Desktop\MGNN-NPI\Select'
file_paths = []
for root, dirs, files in os.walk(folder_path):
    for file in files:
        file_path = os.path.join(root, file)
        file_paths.append(file_path)
for file_path in file_paths:
    graph_data_list = torch.load(file_path)
    device = torch.device('cuda' if torch.cuda.is_available() else 'cuda')
    criterion = torch.nn.BCEWithLogitsLoss()
    def train(model, train_loader, optimizer, criterion):
        model.train()
        total_loss = 0
        for data in train_loader:
            data = data.to(device)
            optimizer.zero_grad()
            out = model(data).squeeze()
            loss = criterion(out, data.y.float().view(-1).to(device))
            loss.backward()
            optimizer.step()
            total_loss += loss.item()
        return total_loss / len(train_loader)
    def validate(model, val_loader, criterion):
        model.eval()
        val_loss = 0
        preds, labels = [], []
        for data in val_loader:
            with torch.no_grad():
                out = model(data).squeeze()
                loss = criterion(out, data.y.float().view(-1).to(device))
                val_loss += loss.item()
                pred = (out > 0).float()
                preds.append(pred)
                labels.append(data.y)
        val_loss /= len(val_loader)
        preds = torch.cat(preds, dim=0)
        labels = torch.cat(labels, dim=0)
        return val_loss, preds, labels
    kf = KFold(n_splits=5, shuffle=True)
    epoch_ = 150
    lr_ = 0.0005
    batch_size_ = 16
    input_dim = int(graph_data_list[0].x.shape[1])
    cumulative_results = {
        'train_loss': [],
        'val_loss': [],
        'val_acc': [],
        'Sp': [],
        'Sn': [],
        'ACC': [],
        'MCC': []
    }
    all_folds_MCC_results = []
    all_folds_ACC_results = []
    all_folds_SN_results = []
    all_folds_SP_results = []
    for fold, (train_idx, val_idx) in enumerate(kf.split(graph_data_list)):
        train_subsampler = SubsetRandomSampler(train_idx)
        val_subsampler = SubsetRandomSampler(val_idx)
        train_loader = DataLoader(graph_data_list, batch_size=batch_size_, sampler=train_subsampler)
        val_loader = DataLoader(graph_data_list, batch_size=batch_size_, sampler=val_subsampler)
        model = GCN(input_dim=input_dim, dropout_rate=0.5).to(device)
        optimizer = torch.optim.Adam(model.parameters(), lr=lr_, weight_decay=5e-4)
        criterion = torch.nn.BCEWithLogitsLoss()
        fold_results = {
            'train_loss': [],
            'val_loss': [],
            'val_acc': [],
            'Sp': [],
            'Sn': [],
            'ACC': [],
            'MCC': []
        }
        for epoch in range(epoch_):
            train_loss = train(model.to(device), train_loader, optimizer, criterion.to(device))
            val_loss, val_preds, val_labels = validate(model.to(device), val_loader, criterion.to(device))
            val_acc = accuracy_score(val_labels.cpu().numpy(), val_preds.cpu().numpy())
            fold_results['train_loss'].append(train_loss)
            fold_results['val_loss'].append(val_loss)
            fold_results['val_acc'].append(val_acc)
            tn, fp, fn, tp = confusion_matrix(val_labels.cpu().numpy(), val_preds.cpu().numpy(), labels=[0, 1]).ravel()
            Sp = tn / (tn + fp)
            Sn = tp / (tp + fn)
            ACC = (tp + tn) / (tp + fp + fn + tn)
            MCC = matthews_corrcoef(val_labels.cpu().numpy(), val_preds.cpu().numpy())
            fold_results['Sp'].append(Sp)
            fold_results['Sn'].append(Sn)
            fold_results['ACC'].append(ACC)
            fold_results['MCC'].append(MCC)
        all_folds_MCC_results.append(fold_results['MCC'])
        all_folds_ACC_results.append(fold_results['ACC'])
        all_folds_SN_results.append(fold_results['Sn'])
        all_folds_SP_results.append(fold_results['Sp'])
    mcc_averages = calculate_average_mcc(all_folds_MCC_results)
    best_epoch_index = mcc_averages.index(max(mcc_averages))
    best_mcc_values = [folds_mcc[best_epoch_index] for folds_mcc in all_folds_MCC_results]
    best_acc_values = [folds_acc[best_epoch_index] for folds_acc in all_folds_ACC_results]
    best_sn_values = [folds_sn[best_epoch_index] for folds_sn in all_folds_SN_results]
    best_sp_values = [folds_sp[best_epoch_index] for folds_sp in all_folds_SP_results]
    average_mcc = sum(best_mcc_values) / len(best_mcc_values)
    average_acc = sum(best_acc_values) / len(best_acc_values)
    average_sn = sum(best_sn_values) / len(best_sn_values)
    average_sp = sum(best_sp_values) / len(best_sp_values)
    print(f"Max MCC in  {best_epoch_index} Epoch, MCC, ACC, SN and SP is：")
    print(f"MCC: {average_mcc}")
    print(f"ACC: {average_acc}")
    print(f"SN: {average_sn}")
    print(f"SP: {average_sp}")
    data_to_save = {
        f"Fold {fold + 1} Results": [
            f"file_path:{file_path.split('_')[-1]}"
            f"Specificity: {average_sp}",
            f"Sensitivity: {average_sn}",
            f"ACC: {average_acc}",
            f"MCC: {average_mcc}"
        ]
    }
    df_to_save = pd.DataFrame([data_to_save])
    try:
        book = load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        startrow = book[f"Fold {fold + 1} Results"].max_row if f"Fold {fold + 1} Results" in book.sheetnames else 0
    except FileNotFoundError:
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        startrow = 0
    df_to_save.to_excel(writer, sheet_name=f"Fold {fold + 1} Results", index=False, header=startrow == 0, startrow=startrow)
    writer.save()
    writer.close()
